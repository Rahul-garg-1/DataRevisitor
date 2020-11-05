import pandas as pd
import requests 
import openpyxl 
import urlparse 
from urllib.request import urljoin 
from bs4 import BeautifulSoup 
from urllib.error import HTTPError
from urllib.error import URLError


# Functions

def tentative_homepage_urls_adder(url,MAX_DEPTH,base_url):
  faculty_keys=['faculty','professor','researcher','doctorate','teacher','lecturer','assistant','people','staff','affiliates']
  pending = list()
  pending.append(url)
  visited=set()
  depth = 0
  ten_fauclty_urls = set()

  while(depth<MAX_DEPTH and len(pending)!=0):
    print("At depth "+ str(depth))
    size=len(pending)
    depth=depth+1

    while(size!=0):
      size=size-1
      link = pending.pop(0)
      if link in visited:
        continue
      visited.add(link)
      print(link)
      
      try:
        page = requests.get(link)
        soup = BeautifulSoup(page.content, 'html.parser')
        all_links = soup.find_all('a')
        ten_fauclty_urls.add(link)
      except HTTPError as e:
        print("404 Not Found")
        continue
      except URLError as e:
        print("The server can't be reached")
        continue
      except:
        print("Some unexpected error occurred")
        continue
      for li in all_links:
        res = False
        for ele in faculty_keys:
          condition = str(li.text)
          if ele in condition.lower():
            res = True
            
        if res and li.get('href')!=None and len(li.get('href'))!=0:
          if "http" not in li.get('href'):
            if li.get('href') not in visited:
              pending.append(base_url+str(li.get('href')))
          else:
            if li.get('href') not in visited:
              pending.append(li.get('href')) 

  return ten_fauclty_urls

def is_faculty_page(url):
  try:
    page = requests.get(url)
    soup = BeautifulSoup(page.content, 'html.parser')
    names_cnt = 0
    all_links = soup.find_all('a')
    for link in all_links:
      result = is_indian_name(link.text)
      if result:
        names_cnt = names_cnt+1
    
    if names_cnt>=7:
      return True
    return False
  except HTTPError as e:
    print("404 Not Found")
    return None
  except URLError as e:
    print("The server can't be reached")
    return None
  except:
    print("Some unexpected error occurred")
    return None



def is_dict_words(text):
  text = text.lower()
  text_list = text.split(" ")
  for ele in text_list:
    if ele in dic_keys:
      return True
  return False


def is_indian_name(full_name):
  full_name = full_name.lower()
  name_list = full_name.split(" ")
  first_name = name_list[0]
  last_name = name_list[-1]
  if len(name_list)==1 and first_name in first_names:
    return True
  if last_name in last_names:
    return True
  return False

def feedback(indian_names):
  print('Predicted Indian Names(Press y if correct n if incorrect):')
  for name in indian_names:
    print(name+' (y/n): ')
    option=input()
    if option.lower()=='y':
      continue
    name_list = name.split(" ")
    name_list = [ele.lower() for ele in name_list]
    for every_name in name_list:
      if is_indian_name(every_name):
        print("Is "+every_name+' Indian name(y/n): ')
        option1 = input()
        if option1.lower()=='n':
          first_names.discard(every_name)
          last_names.discard(every_name)
          print('Names Database updated')



# Reading the datasets

df = pd.read_excel("/gdrive/My Drive/datasets/surnames_indian.xlsx")
last_names = df['Names'].tolist()
last_names = [ele.lower() for ele in last_names]
last_names = set(last_names)

df1 = pd.read_excel("/gdrive/My Drive/datasets/indian_first_names.xlsx")
first_names = df1['Names'].tolist()
first_names = [ele.lower() for ele in first_names]
first_names = set(first_names)

dataset = pd.read_excel("/gdrive/My Drive/datasets/faculty_urls.xlsx")
original_urls = dataset['url'].tolist()
# URL = "https://pec.ac.in"
# tentative_homepage_urls_adder(URL)
tentative_urls=set()


for url in original_urls:
  first_id = url.find('/')
  second_id = url.find('/',first_id+1)
  third_id = url.find('/',second_id+1)
  url = url[:third_id]
  return_list = tentative_homepage_urls_adder(url,3,url)
  tentative_urls = list(tentative_urls)
  tentative_urls.extend(return_list)
  tentative_urls = set(tentative_urls)

for url in original_urls:
  link = url[:url.rindex('/')]
  first_id = url.find('/')
  second_id = url.find('/',first_id+1)
  third_id = url.find('/',second_id+1)
  base_url = url[:third_id]
  return_list = tentative_homepage_urls_adder(url,2,base_url)
  tentative_urls = list(tentative_urls)
  tentative_urls.extend(return_list)
  tentative_urls = set(tentative_urls)

print(tentative_urls)
faculty_homepage_urls = set()

for url in tentative_urls:
  if is_faculty_page(url):
    faculty_homepage_urls.add(url)


dictionary = pd.read_excel("/gdrive/My Drive/datasets/dictionary.xlsx")
dic_keys = set(dictionary['Dictionary'].tolist())


# NEW faculties addition

indian_names_dict = dict()

for link in faculty_homepage_urls:
  try:
    page = requests.get(link)
    soup = BeautifulSoup(page.content, 'html.parser')
    all_links = soup.find_all('a')
  except HTTPError as e:
    print("404 Not Found")
    continue
  except URLError as e:
    print("The server can't be reached")
    continue
  except:
    print("Some unexpected error occurred")
    continue
  first_id = link.find('/')
  second_id = link.find('/',first_id+1)
  third_id = link.find('/',second_id+1)
  base_url = link[:third_id]
  for every_link in all_links:
    if is_indian_name(every_link.text.strip()):
      absolute_url = every_link.get('href')
      if "http" not in every_link.get('href'):
        absolute_url = base_url + every_link.get('href')
      if is_dict_words(every_link.text.strip()) == False:
        indian_names_dict[absolute_url] = every_link.text.strip()

new_faculty_urls = []
for key in indian_names_dict.keys():
  if key not in original_urls:
    new_faculty_urls.append(key)

updated_indian_faculty_dict = dict()
for key in new_faculty_urls:
  updated_indian_faculty_dict[key] = indian_names_dict[key]

for key, value in updated_indian_faculty_dict.items():
  print(key+" -- "+value)


final_new_faculty_urls = []
for key in new_faculty_urls:
  final_new_faculty_urls.append(key)




wb = openpyxl.Workbook() 
sheet = wb.active 
c1 = sheet.cell(row=1,column=1)
c1.value = "Name"
c2 = sheet.cell(row=1,column=2)
c2.value = "URL"
r = 2
for key,value in updated_indian_faculty_dict.items():
  cell1 = sheet.cell(r,column=1)
  cell1.value = value
  cell2 = sheet.cell(r,column=2)
  cell2.value = key
  r=r+1


wb.save("/gdrive/My Drive/datasets/new_named_urls.xlsx")